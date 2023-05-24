import csv
import pprint
import random

from openpyxl import load_workbook

import requests

import re

import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer

from nltk.stem import PorterStemmer

from nltk.corpus import words

import nltk

import random
import string

nltk.download('words')


def gen_rand_letter_number(k=2):
    characters = string.ascii_letters + string.digits
    random_string = ''.join(random.choices(characters, k=k))
    return random_string


def replace_punctuation_and_whitespace(string):
    stripped = string.strip()

    # Replace punctuation and whitespace with underscores
    replaced = re.sub(r'\W+', '_', stripped)

    # Remove consecutive underscores
    final_result = re.sub(r'_{2,}', '_', replaced)

    # trim leading and trailing whitespace

    return final_result


url = 'https://github.com/GenomicsStandardsConsortium/mixs/raw/main/mixs/excel/mixs_v6.xlsx'
file_path = 'mixs_v6.xlsx'

# Download the Excel file
response = requests.get(url)
with open(file_path, 'wb') as file:
    file.write(response.content)

# Load the Excel file
workbook = load_workbook(file_path)

# Read the "MIxS" sheet
sheet_mixs = workbook['MIxS']
headers_mixs = [cell.value for cell in sheet_mixs[1]]

data_mixs = [dict(zip(headers_mixs, [cell.value for cell in row])) for row in sheet_mixs.iter_rows(min_row=2)]

# Read the "environmental_packages" sheet
sheet_env_packages = workbook['environmental_packages']

# Optionally, close the workbook
workbook.close()

headers_env_packages = [cell.value for cell in sheet_env_packages[1]]

data_env_packages = [dict(zip(headers_env_packages, [cell.value for cell in row])) for row in
                     sheet_env_packages.iter_rows(min_row=2)]

mixs_only = list(set(headers_mixs) - set(headers_env_packages))

env_packages_only = list(set(headers_env_packages) - set(headers_mixs))

# Use data_mixs and data_env_packages for further analysis or processing

deletables = [
    " ",
    "Requirement",
    'migs_ba',
    'migs_eu'
    'migs_org',
    'migs_pl',
    'migs_vi',
    'mimag',
    'mimarks_c',
    'mimarks_s',
    'mims',
    'misag',
    'miuvig',
    None,
]

for row in data_mixs:
    if 'Occurence' in row:
        row['Occurrence'] = row.pop('Occurence')
    if 'Item (rdfs:label)' in row:
        row['Item'] = row.pop('Item (rdfs:label)')
    for d in deletables:
        if d in row:
            del row[d]

for row in data_env_packages:
    if 'Package item' in row:
        row['Item'] = row.pop('Package item')
    for d in deletables:
        if d in row:
            del row[d]

combined_sheets = data_mixs + data_env_packages

for row in combined_sheets:
    if 'Environmental package' in row:
        differentiator = replace_punctuation_and_whitespace(row['Environmental package'])
    elif 'Section' in row:
        differentiator = replace_punctuation_and_whitespace(row['Section'])
    else:
        differentiator = ''

    # random_number = random.randint(0, 999)
    # padded_number = str(random_number).zfill(3)

    gen_rand_letter_number(k=2)

    # Structured comment name

    # row[
    #     'unique_id'] = f"{row['MIXS ID']}:{replace_punctuation_and_whitespace(row['Item'])}:{differentiator}:{padded_number}"

    row[
        'unique_id'] = f"{replace_punctuation_and_whitespace(row['Structured comment name'])}:{differentiator}:{row['MIXS ID']}:{gen_rand_letter_number(k=2)}"

# # Optionally, remove the downloaded file
# # import os
# # os.remove(file_path)

# Convert list of dictionaries to a DataFrame
df = pd.DataFrame(combined_sheets)

df.to_csv('mixs_envo_struct_knowl_extraction.tsv', index=False, sep='\t')

# Create an instance of CountVectorizer with stop word removal, custom tokenizer, and modified token pattern
vectorizer = CountVectorizer(
    stop_words='english',
    tokenizer=lambda text: text.split(),
    token_pattern=r'[a-zA-Z]{2,}',
    # max_features=2500,
    # max_df=0.01
)

# Fit the vectorizer on the Definition column and transform it into a document-term matrix
matrix = vectorizer.fit_transform(df['Definition'])

# Get the feature names (terms)
feature_names = vectorizer.get_feature_names_out()

# Load a dictionary of valid words (e.g., from NLTK's word corpus)

valid_words = set(words.words())

# Filter the feature names to include only dictionary words
dictionary_feature_names = [word for word in feature_names if word in valid_words]

# Apply stemming to the dictionary feature names using PorterStemmer
stemmer = PorterStemmer()
stemmed_feature_names = [stemmer.stem(word) for word in dictionary_feature_names]

# Filter the document-term matrix to include only the columns corresponding to the dictionary feature names
filtered_matrix = matrix[:, [vectorizer.vocabulary_[word] for word in dictionary_feature_names]]

# Create a new DataFrame for the filtered document-term matrix with the stemmed dictionary feature names
dtm_df = pd.DataFrame(filtered_matrix.toarray(), columns=stemmed_feature_names, index=df['unique_id'])

# Print the resulting DataFrame
print(dtm_df)
print("Feature names:", dictionary_feature_names)

dtm_df.to_csv('dtm.csv', index=True, sep='\t')
